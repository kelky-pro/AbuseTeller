import asyncio
import csv
import json
import logging
import os
import random
import re
import time
from io import BytesIO, StringIO
from math import ceil

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.dimensions import ColumnDimension
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import ApplicationBuilder, CallbackQueryHandler, CommandHandler, ContextTypes, MessageHandler, filters
from tenacity import retry, stop_after_attempt, wait_fixed, retry_if_exception_type

# Setup logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Configuration via environment variables
CONFIG = {
    "TELEGRAM_BOT_TOKEN": os.getenv("TELEGRAM_BOT_TOKEN", ""),
    "ABUSEIPDB_API_KEY_1": os.getenv("ABUSEIPDB_API_KEY_1", "eca572ae930dd61ded6eb59112d8bb15ea657fb34a069ec89543fa9a0d6e47ecd5bac8457566bdf3"),
    "ABUSEIPDB_API_KEY_2": os.getenv("ABUSEIPDB_API_KEY_2", "399456c14ee1b25c6cc9218a3257eeee373259c9aa813e0cddbe3ba296cbd651a975e70b11b8de33"),
    "MAX_IPS": int(os.getenv("MAX_IPS", 100)),
    "BATCH_SIZE": int(os.getenv("BATCH_SIZE", 10)),
    "API_RETRIES": int(os.getenv("API_RETRIES", 3)),
    "WELL_KNOWN_ISPS": json.loads(os.getenv("WELL_KNOWN_ISPS", '{"Microsoft": ["microsoft", "azure"], "Google": ["google"], "GitHub": ["github"]}')),
    "ETHIOPIAN_ISP_KEYWORDS": os.getenv("ETHIOPIAN_ISP_KEYWORDS", "ethio,ethiopian").lower().split(","),
    "REQUESTS_PER_DAY_PER_KEY": 1000,
    "THREAT_THRESHOLD": int(os.getenv("THREAT_THRESHOLD", 50)),
    "OUTPUT_FORMAT": os.getenv("OUTPUT_FORMAT", "excel").lower(),
    "ALLOW_FILE_INPUT": os.getenv("ALLOW_FILE_INPUT", "true").lower() == "true",
    "FALLBACK_API": os.getenv("FALLBACK_API", "none").lower(),
    "ANIMATION_STYLE": os.getenv("ANIMATION_STYLE", "spinner").lower(),
    "EXCEL_STYLE": json.loads(os.getenv("EXCEL_STYLE", '{"header_color": "008000", "summary_color": "FFFF00", "high_abuse_color": "FF0000"}')),
}

# Track API usage for each key
API_USAGE = {
    "key1": {"count": 0, "last_reset": time.time()},
    "key2": {"count": 0, "last_reset": time.time()}
}

# English messages
MESSAGES = {
    "welcome": "Welcome to @kelbudget_bot! 💻 Lookup up to {max_ips} IPs via text or .txt file and get results in {format} format.\nUse /lookup, /stats, or the button below:",
    "send_ips": "Send IP addresses, one per line (e.g., '8.8.8.8\n1.1.1.1'), or upload a .txt file:",
    "stats": "📊 API Usage:\nKey 1: {key1_count}/{max_requests} requests\nKey 2: {key2_count}/{max_requests} requests\nRemaining: {total_remaining} requests today",
    "no_api_key": "❌ Bot configuration error: API key(s) missing.",
    "no_token": "❌ Bot configuration error: Telegram token missing.",
    "invalid_ips": "❌ Invalid IP(s): {ips}. Try again.",
    "no_ips": "❌ No IP addresses provided. Try again.",
    "invalid_file": "❌ Invalid file. Please upload a .txt file with IPs, one per line.",
    "file_too_large": "❌ File too large. Please upload a file smaller than 1MB.",
    "too_many_ips": "❌ Too many IPs. Maximum allowed is {max_ips}.",
    "api_exhausted": "❌ Both API keys have reached their daily limit. Try again tomorrow.",
    "processing": "🔄 Processing {num_ips} IP(s)... {progress}% done\n{animation}",
    "success": "✅ Results for {num_ips} IP(s) in {format} format.",
    "error": "❌ Error: {error}. Try again.",
    "use_lookup": "Use /lookup or the button to check IPs! 💻",
}

# Animation styles
ANIMATION_STYLES = {
    "spinner": ["⏳ [  ]", "⌛ [█ ]", "⏰ [██ ]", "⌛ [███]"],
    "bar": ["[    ]", "[█   ]", "[██  ]", "[███ ]", "[████]"]
}

# Validate IP address
def is_valid_ip(ip):
    ipv4_pattern = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
    ipv6_pattern = r'^([0-9a-fA-F]{1,4}:){7,7}[0-9a-fA-F]{1,4}|([0-9a-fA-F]{1,4}:){1,7}:|([0-9a-fA-F]{1,4}:){1,6}:[0-9a-fA-F]{1,4}|([0-9a-fA-F]{1,4}:){1,5}(:[0-9a-fA-F]{1,4}){1,2}|([0-9a-fA-F]{1,4}:){1,4}(:[0-9a-fA-F]{1,4}){1,3}|([0-9a-fA-F]{1,4}:){1,3}(:[0-9a-fA-F]{1,4}){1,4}|([0-9a-fA-F]{1,4}:){1,2}(:[0-9a-fA-F]{1,4}){1,5}|[0-9a-fA-F]{1,4}:((:[0-9a-fA-F]{1,4}){1,6})|:((:[0-9a-fA-F]{1,4}){1,7}|:)$'
    return bool(re.match(ipv4_pattern, ip) or re.match(ipv6_pattern, ip))

# Query AbuseIPDB API
@retry(
    stop=stop_after_attempt(CONFIG["API_RETRIES"]),
    wait=wait_fixed(2),
    retry=retry_if_exception_type(requests.RequestException),
    before_sleep=lambda retry_state: logging.debug(f"Retrying API call for IP {retry_state.args[0]}: attempt {retry_state.attempt_number}")
)
def query_abuseipdb(ip, api_key, key_id):
    global API_USAGE
    current_time = time.time()
    if current_time - API_USAGE[key_id]["last_reset"] > 86400:
        API_USAGE[key_id]["count"] = 0
        API_USAGE[key_id]["last_reset"] = current_time
    if API_USAGE[key_id]["count"] >= CONFIG["REQUESTS_PER_DAY_PER_KEY"]:
        raise ValueError(f"API request limit reached for {key_id}")
    API_USAGE[key_id]["count"] += 1
    logging.debug(f"Using {key_id}: {API_USAGE[key_id]['count']}/{CONFIG['REQUESTS_PER_DAY_PER_KEY']} requests")
    url = "https://api.abuseipdb.com/api/v2/check"
    headers = {"Key": api_key, "Accept": "application/json"}
    params = {"ipAddress": ip, "maxAgeInDays": 90}
    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()
    data = response.json().get('data', {})
    return {
        'ip': data.get('ipAddress', ip),
        'isp': data.get('isp', 'N/A'),
        'usage_type': data.get('usageType', 'N/A'),
        'domain': data.get('domain', 'N/A'),
        'abuse_score': data.get('abuseConfidenceScore', 0)
    }

# Query ip-api.com as fallback
def query_ip_api(ip):
    url = f"http://ip-api.com/json/{ip}"
    try:
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        data = response.json()
        if data.get("status") != "success":
            raise ValueError("ip-api.com query failed")
        return {
            'ip': ip,
            'isp': data.get('isp', 'N/A'),
            'usage_type': data.get('org', 'N/A'),
            'domain': data.get('as', 'N/A').split()[1] if len(data.get('as', '').split()) > 1 else 'N/A',
            'abuse_score': 0  # ip-api.com doesn't provide abuse score
        }
    except Exception as e:
        logging.error(f"ip-api.com query failed for IP {ip}: {str(e)}")
        return None

# Categorize ISPs (for sorting only)
def categorize_isp(result):
    if not result:
        return None, "Others"
    isp = result['isp'].lower()
    domain = result['domain'].lower() if result['domain'] else ''
    country = result.get('country', 'N/A').lower()
    category = "Others"
    for isp_name, keywords in CONFIG["WELL_KNOWN_ISPS"].items():
        if any(keyword.lower() in isp or keyword.lower() in domain for keyword in keywords):
            category = isp_name
            break
    if country == 'ethiopia' or any(keyword in isp for keyword in CONFIG["ETHIOPIAN_ISP_KEYWORDS"]):
        category = "Ethiopian"
    return result, category

# Generate Excel file
def generate_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "IP Lookup Results"
    ws.protection = None  # Explicitly disable sheet protection
    headers = ["IP Address", "ISP", "Usage Type", "Domain", "Abuse Score (%)"]
    
    # Calculate threat summary
    high_risk_count = sum(1 for result, _ in results if result and result['abuse_score'] > CONFIG["THREAT_THRESHOLD"])
    ws.append(["Summary", f"{high_risk_count} IPs with abuse score >{CONFIG['THREAT_THRESHOLD']}%", "", "", ""])
    
    # Add headers
    ws.append(headers)

    # Styling
    header_font = Font(bold=True)
    summary_font = Font(bold=True, italic=True)
    header_fill = PatternFill(start_color=CONFIG["EXCEL_STYLE"].get("header_color", "008000"), end_color=CONFIG["EXCEL_STYLE"].get("header_color", "008000"), fill_type="solid")
    summary_fill = PatternFill(start_color=CONFIG["EXCEL_STYLE"].get("summary_color", "FFFF00"), end_color=CONFIG["EXCEL_STYLE"].get("summary_color", "FFFF00"), fill_type="solid")
    high_abuse_fill = PatternFill(start_color=CONFIG["EXCEL_STYLE"].get("high_abuse_color", "FF0000"), end_color=CONFIG["EXCEL_STYLE"].get("high_abuse_color", "FF0000"), fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Style summary row
    for cell in ws[1]:
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left')

    # Style header row
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Append data
    for result, _ in results:
        if result:
            ws.append([
                result['ip'], result['isp'], result['usage_type'],
                result['domain'], result['abuse_score']
            ])
        else:
            ws.append(["Failed", "N/A", "N/A", "N/A", 0])
        # Apply borders and conditional formatting
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='left')
        if result and result['abuse_score'] > CONFIG["THREAT_THRESHOLD"]:
            ws.cell(row=ws.max_row, column=5).fill = high_abuse_fill

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Generate CSV file
def generate_csv(results):
    output = StringIO()
    writer = csv.writer(output)
    headers = ["IP Address", "ISP", "Usage Type", "Domain", "Abuse Score (%)"]
    
    # Write threat summary
    high_risk_count = sum(1 for result, _ in results if result and result['abuse_score'] > CONFIG["THREAT_THRESHOLD"])
    writer.writerow(["Summary", f"{high_risk_count} IPs with abuse score >{CONFIG['THREAT_THRESHOLD']}%", "", "", ""])
    writer.writerow([])  # Blank row
    writer.writerow(headers)

    # Write data
    for result, _ in results:
        if result:
            writer.writerow([
                result['ip'], result['isp'], result['usage_type'],
                result['domain'], result['abuse_score']
            ])
        else:
            writer.writerow(["Failed", "N/A", "N/A", "N/A", 0])

    output.seek(0)
    return BytesIO(output.getvalue().encode('utf-8'))

# Command: /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    logging.debug(f"Received /start from user_id: {user_id}")
    keyboard = [[InlineKeyboardButton("Lookup IPs", callback_data="lookup_ips")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        MESSAGES["welcome"].format(max_ips=CONFIG["MAX_IPS"], format=CONFIG["OUTPUT_FORMAT"].upper()),
        reply_markup=reply_markup
    )

# Command: /stats
async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    logging.debug(f"Received /stats from user_id: {user_id}")
    total_remaining = (CONFIG["REQUESTS_PER_DAY_PER_KEY"] - API_USAGE["key1"]["count"]) + (CONFIG["REQUESTS_PER_DAY_PER_KEY"] - API_USAGE["key2"]["count"])
    await update.message.reply_text(
        MESSAGES["stats"].format(
            key1_count=API_USAGE["key1"]["count"],
            key2_count=API_USAGE["key2"]["count"],
            max_requests=CONFIG["REQUESTS_PER_DAY_PER_KEY"],
            total_remaining=total_remaining
        )
    )

# Command: /lookup
async def lookup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    logging.debug(f"Received /lookup from user_id: {user_id}")
    context.user_data['awaiting_ips'] = True
    await update.message.reply_text(MESSAGES["send_ips"])

# Callback query handler
async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "lookup_ips":
        context.user_data['awaiting_ips'] = True
        await query.message.reply_text(MESSAGES["send_ips"])

# Handle text input
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get('awaiting_ips', False):
        ips = [ip.strip() for ip in update.message.text.splitlines() if ip.strip()]
        context.user_data['awaiting_ips'] = False
        await process_ips(ips, update, context)
    else:
        await update.message.reply_text(MESSAGES["use_lookup"])

# Handle file input
async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not CONFIG["ALLOW_FILE_INPUT"]:
        await update.message.reply_text("❌ File uploads are disabled.")
        return
    if not update.message.document or not update.message.document.file_name.endswith('.txt'):
        await update.message.reply_text(MESSAGES["invalid_file"])
        return
    try:
        file = await update.message.document.get_file()
        if file.file_size > 1024 * 1024:  # Limit to 1MB
            await update.message.reply_text(MESSAGES["file_too_large"])
            return
        file_content = await file.download_as_bytearray()
        ips = [ip.strip() for ip in file_content.decode('utf-8').splitlines() if ip.strip()]
        if len(ips) > CONFIG["MAX_IPS"]:
            await update.message.reply_text(MESSAGES["too_many_ips"].format(max_ips=CONFIG["MAX_IPS"]))
            return
        await process_ips(ips, update, context)
    except Exception as e:
        logging.error(f"Error processing file: {str(e)}")
        await update.message.reply_text(MESSAGES["error"].format(error=str(e)))

# Process IPs (text or file)
async def process_ips(ips, update, context):
    if not (CONFIG["ABUSEIPDB_API_KEY_1"] and CONFIG["ABUSEIPDB_API_KEY_2"]):
        await update.message.reply_text(MESSAGES["no_api_key"])
        return
    if not CONFIG["TELEGRAM_BOT_TOKEN"]:
        await update.message.reply_text(MESSAGES["no_token"])
        return
    invalid_ips = [ip for ip in ips if not is_valid_ip(ip)]
    if invalid_ips:
        await update.message.reply_text(MESSAGES["invalid_ips"].format(ips=', '.join(invalid_ips[:5])))
        return
    if not ips:
        await update.message.reply_text(MESSAGES["no_ips"])
        return
    # Check API key exhaustion
    if API_USAGE["key1"]["count"] >= CONFIG["REQUESTS_PER_DAY_PER_KEY"] and API_USAGE["key2"]["count"] >= CONFIG["REQUESTS_PER_DAY_PER_KEY"]:
        await update.message.reply_text(MESSAGES["api_exhausted"])
        return
    try:
        # Send initial processing message
        animation_frames = ANIMATION_STYLES.get(CONFIG["ANIMATION_STYLE"], ANIMATION_STYLES["spinner"])
        status_message = await update.message.reply_text(
            MESSAGES["processing"].format(num_ips=len(ips), progress=0, animation=animation_frames[0])
        )
        batch_size = min(CONFIG["BATCH_SIZE"], len(ips))
        results = []
        total_batches = ceil(len(ips) / batch_size)
        for i in range(0, len(ips), batch_size):
            batch = ips[i:i + batch_size]
            batch_results = []
            for ip in batch:
                result = None
                try:
                    # Try AbuseIPDB with random key
                    key_id, api_key = random.choice([("key1", CONFIG["ABUSEIPDB_API_KEY_1"]), ("key2", CONFIG["ABUSEIPDB_API_KEY_2"])])
                    result = query_abuseipdb(ip, api_key, key_id)
                except Exception as e:
                    logging.error(f"AbuseIPDB query failed for IP {ip}: {str(e)}")
                    # Try fallback API if enabled
                    if CONFIG["FALLBACK_API"] == "ip-api":
                        result = query_ip_api(ip)
                if result:
                    batch_results.append(categorize_isp(result))
                else:
                    batch_results.append((None, "Others"))
            results.extend(batch_results)
            # Update animation
            progress = int(((i + batch_size) / len(ips)) * 100)
            frame = animation_frames[(i // batch_size) % len(animation_frames)]
            try:
                await context.bot.edit_message_text(
                    chat_id=status_message.chat_id,
                    message_id=status_message.message_id,
                    text=MESSAGES["processing"].format(num_ips=len(ips), progress=progress, animation=frame)
                )
            except Exception as e:
                logging.debug(f"Failed to update animation: {str(e)}")
                status_message = await update.message.reply_text(
                    MESSAGES["processing"].format(num_ips=len(ips), progress=progress, animation=frame)
                )
            await asyncio.sleep(random.uniform(0.5, 1.5))  # Random delay to avoid rate limits
        # Sort results by category
        results.sort(key=lambda x: x[1])
        # Generate output file
        output = None
        file_name = None
        if CONFIG["OUTPUT_FORMAT"] == "excel":
            output = generate_excel(results)
            file_name = "ip_lookup_results.xlsx"
        elif CONFIG["OUTPUT_FORMAT"] == "csv":
            output = generate_csv(results)
            file_name = "ip_lookup_results.csv"
        # Send results
        await context.bot.send_document(
            chat_id=update.message.chat_id,
            document=output,
            filename=file_name,
            caption=MESSAGES["success"].format(num_ips=len(ips), format=CONFIG["OUTPUT_FORMAT"].upper())
        )
    except Exception as e:
        logging.error(f"Error processing IPs: {str(e)}")
        await update.message.reply_text(MESSAGES["error"].format(error=str(e)))

# Main application setup
async def main():
    if not CONFIG["TELEGRAM_BOT_TOKEN"]:
        logging.error("Telegram bot token is missing.")
        return
    application = ApplicationBuilder().token(CONFIG["TELEGRAM_BOT_TOKEN"]).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("stats", stats))
    application.add_handler(CommandHandler("lookup", lookup))
    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_file))
    await application.run_polling()

if __name__ == "__main__":
    asyncio.run(main())
